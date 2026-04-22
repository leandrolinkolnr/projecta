import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.title("Relatório de Vendas - Atacado")

# =========================
# PROCESSAMENTO DOS DADOS
# =========================

data = pd.read_excel("tempiii.xlsx")

# Agrupando meses de acordo com o RCA
data = data.pivot_table(
    index=['CODUSUR', 'NOME'],
    columns='MES',
    values='VLTOTAL',
    aggfunc='sum'
).reset_index()

data.columns.name = None

# Criando media 2 meses
data['MEDIA_2M'] = data[['01/2026', '02/2026']].mean(axis=1)
data = data[['CODUSUR', 'NOME', '01/2026', '02/2026', 'MEDIA_2M', '03/2026']]

# Comparando o mes 3 com a media dos 2 primeiros
data['DESVIO_%'] = ((data['03/2026'] - data['MEDIA_2M']) / data['MEDIA_2M'] * 100).round(1)

# Selecionando se Desvio < -5%
data = data[data['DESVIO_%'] < -5]

# Ordenando
data = data.sort_values(by='DESVIO_%', ascending=True)

# Formatando
cols_moeda = ['01/2026', '02/2026', 'MEDIA_2M', '03/2026']

for col in cols_moeda:
    data[col] = data[col].round(0).astype(int)
    data[col] = data[col].apply(lambda x: f'R$ {x:,}'.replace(',', '.'))

# =========================
# EXPORTAÇÃO EXCEL EM MEMÓRIA
# =========================

output = BytesIO()
arquivo = "resultadoz.xlsx"

data.to_excel(output, index=False, engine='openpyxl')
output.seek(0)

# Ajuste de largura no Excel (≈185px = 26.4)
with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
    data.to_excel(writer, index=False, sheet_name='Sheet1')
    writer.book.save(arquivo)

wb = load_workbook(arquivo)
ws = wb.active

largura_padrao = 26.4

for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = largura_padrao

wb.save(arquivo)

# =========================
# BOTÃO DE DOWNLOAD
# =========================

with open(arquivo, "rb") as f:
    st.download_button(
        label="📥 Baixar relatório",
        data=f,
        file_name="resultadoz.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
